from decimal import Decimal, InvalidOperation
from io import BytesIO
import json
import re
import subprocess

from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Count, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from openpyxl import Workbook, load_workbook

from modules.facility.views import (
    auto_debit_claim_page,
    auto_debit_claim_voucher_pdf,
)

from .models import BUDGET_PAGE_CHOICES, BudgetChangeLog, BudgetItem

FIELDS = [
    ('category', '分類'),
    ('budget_code', '2026預算代號'),
    ('ministry', '事工'),
    ('annual_goal', '年度目標'),
    ('strategy_plan', '策略&執行計畫'),
    ('activity_budget', '活動與預算'),
    ('lead_pastor', '主責牧者'),
    ('budget_2026', '2026預算'),
    ('accounting_subject', '會計科目'),
]
FORM_FIELDS = FIELDS[:8] + [('used_amount', '已使用金額')] + FIELDS[8:]

HEADER_ALIASES = {
    '分類': 'category',
    '2026預算代號': 'budget_code',
    '事工': 'ministry',
    '年度目標': 'annual_goal',
    '策略&執行計畫': 'strategy_plan',
    '策略＆執行計畫': 'strategy_plan',
    '活動與預算': 'activity_budget',
    '主責牧者': 'lead_pastor',
    '2026預算': 'budget_2026',
    '會計科目': 'accounting_subject',
    '已使用金額': 'used_amount',
    '使用金額': 'used_amount',
    '美美紀錄': 'used_amount',
    '所屬分頁': 'page_group',
}

TRACKED_FIELDS = ['page_group'] + [field for field, _label in FORM_FIELDS]

BUDGET_PAGE_GROUPS = [
    {
        'slug': 'staff-special-reserve',
        'name': '同工-特別-預備',
        'categories': ('人事薪資', '全職同工', '董執團隊', '特別計畫', '預備金', '國度基金'),
    },
    {'slug': 'administration', 'name': '行政', 'categories': ('行政部',)},
    {'slug': 'worship', 'name': '崇拜', 'categories': ('崇拜部',)},
    {'slug': 'education', 'name': '教育', 'categories': ('教育部',)},
    {'slug': 'mission', 'name': '宣教', 'categories': ('宣教部',)},
    {'slug': 'care', 'name': '關懷', 'categories': ('關懷部',)},
    {'slug': 'counseling', 'name': '輔導', 'categories': ('輔導部',)},
    {'slug': 'technology', 'name': '科技', 'categories': ('科技服務部',)},
    {
        'slug': 'gospel',
        'name': '福音',
        'categories': ('重修舊好志工團', '伯利恆糧食之家\n(BLH)'),
    },
    {
        'slug': 'pastoral-one',
        'name': '牧區一',
        'categories': (
            '牧區處\nPA', '二魚', '多多牧區', '清一牧區', '清二牧區', '幸福牧區',
            '百合A區', '百合B區', '百合C區', '橄欖樹牧區', '青草地牧區',
            '青橄欖', '房角石牧區',
        ),
    },
    {
        'slug': 'pastoral-two',
        'name': '牧區二',
        'categories': (
            'young牧區', '兒童牧區', '百基拉牧區', '三一牧區', '蒙愛查經團契',
            '弟兄會', '加樂團契', '蒙恩團契',
        ),
    },
]


def _clean(value):
    if value is None:
        return ''
    if isinstance(value, str):
        return value.strip()
    return value


def _page_group_for_category(category):
    for page in BUDGET_PAGE_GROUPS:
        if category in page['categories']:
            return page['slug']
    return BUDGET_PAGE_GROUPS[0]['slug']


def _page_group_from_value(value, category=''):
    cleaned_value = str(_clean(value) or '')
    page_by_value = {
        choice_value: choice_value
        for choice_value, _choice_label in BUDGET_PAGE_CHOICES
    }
    page_by_value.update({
        choice_label: choice_value
        for choice_value, choice_label in BUDGET_PAGE_CHOICES
    })
    return page_by_value.get(cleaned_value, _page_group_for_category(category))


def _decimal_or_none(value):
    value = _clean(value)
    if value == '':
        return None
    if isinstance(value, str):
        value = value.replace(',', '').replace('$', '').strip()
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _payload_from_post(post, current_page_group=None):
    payload = {}
    for field, _label in FORM_FIELDS:
        if field in ('budget_2026', 'used_amount'):
            payload[field] = _decimal_or_none(post.get(field))
        else:
            payload[field] = (post.get(field) or '').strip()
    posted_page_group = post.get('page_group')
    if posted_page_group is None and current_page_group is not None:
        posted_page_group = current_page_group
    payload['page_group'] = _page_group_from_value(posted_page_group, payload.get('category', ''))
    return payload


def _snapshot(item):
    if item is None:
        return None
    data = {}
    for field in TRACKED_FIELDS:
        value = getattr(item, field)
        data[field] = str(value) if isinstance(value, Decimal) else value
    return data


def _client_ip(request):
    forwarded = request.META.get('HTTP_X_FORWARDED_FOR', '')
    if forwarded:
        return forwarded.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR')


def _client_mac(ip_address):
    if not ip_address:
        return ''
    commands = [
        ['ip', 'neigh', 'show', ip_address],
        ['arp', '-n', ip_address],
    ]
    for command in commands:
        try:
            result = subprocess.run(command, capture_output=True, text=True, timeout=2)
        except Exception:
            continue
        match = re.search(r'([0-9a-fA-F]{2}(?::[0-9a-fA-F]{2}){5})', result.stdout)
        if match:
            return match.group(1).lower()
    return ''


def _log_change(request, action, item=None, before=None, after=None):
    ip_address = _client_ip(request)
    user = request.user if request.user.is_authenticated else None
    BudgetChangeLog.objects.create(
        budget_item=item,
        action=action,
        before_data=before,
        after_data=after,
        changed_by=user,
        changed_by_code=getattr(user, 'username', '') if user else '',
        ip_address=ip_address or None,
        mac_address=_client_mac(ip_address),
    )


def _sheet_for_import(workbook):
    if '合併資料' in workbook.sheetnames:
        return workbook['合併資料']
    return workbook.active


def _header_map(sheet):
    mapping = {}
    for col in range(1, sheet.max_column + 1):
        header = _clean(sheet.cell(1, col).value)
        if header in HEADER_ALIASES:
            mapping[HEADER_ALIASES[header]] = col
    required = [field for field, _label in FIELDS]
    if all(field in mapping for field in required):
        return mapping

    # Prepared 2026 budget sheet: A=部門牧區, B-J=target fields, K=used amount.
    fallback = {field: index for (field, _label), index in zip(FIELDS, range(2, 11))}
    fallback['used_amount'] = 11
    return fallback


def import_budget_items(file_obj, request=None):
    workbook = load_workbook(file_obj, data_only=True)
    sheet = _sheet_for_import(workbook)
    mapping = _header_map(sheet)
    items = []
    for row_idx in range(2, sheet.max_row + 1):
        row_data = {}
        for field, _label in FORM_FIELDS:
            col = mapping.get(field)
            value = sheet.cell(row_idx, col).value if col else None
            row_data[field] = _decimal_or_none(value) if field in ('budget_2026', 'used_amount') else str(_clean(value) or '')
        page_group_col = mapping.get('page_group')
        page_group_value = sheet.cell(row_idx, page_group_col).value if page_group_col else None
        row_data['page_group'] = _page_group_from_value(page_group_value, row_data['category'])
        if not any(
            row_data[field] not in ('', None)
            for field, _label in FORM_FIELDS
            if field not in ('budget_2026', 'used_amount')
        ) and row_data['budget_2026'] is None and row_data['used_amount'] is None:
            continue
        items.append(BudgetItem(**row_data))

    with transaction.atomic():
        before_count = BudgetItem.objects.count()
        BudgetItem.objects.all().delete()
        if items:
            BudgetItem.objects.bulk_create(items, batch_size=500)
        if request is not None:
            _log_change(request, 'import', None, {'count': before_count}, {'count': len(items)})
    return len(items)


@login_required
def budget_list(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'create':
            item = BudgetItem.objects.create(**_payload_from_post(request.POST))
            _log_change(request, 'create', item, None, _snapshot(item))
            return redirect(request.POST.get('next') or 'budget-list')
        if action == 'update':
            item = get_object_or_404(BudgetItem, pk=request.POST.get('item_id'))
            before = _snapshot(item)
            for field, value in _payload_from_post(request.POST, item.page_group).items():
                setattr(item, field, value)
            item.save()
            _log_change(request, 'update', item, before, _snapshot(item))
            return redirect(request.POST.get('next') or 'budget-list')
        if action == 'import' and request.FILES.get('file'):
            import_budget_items(request.FILES['file'], request)
            return redirect('budget-list')
        return redirect('budget-list')

    query = (request.GET.get('q') or '').strip()
    page_counts = dict(
        BudgetItem.objects
        .values_list('page_group')
        .annotate(item_count=Count('id'))
        .order_by()
    )
    valid_page_slugs = {page['slug'] for page in BUDGET_PAGE_GROUPS}
    unassigned_count = sum(
        count for slug, count in page_counts.items()
        if slug not in valid_page_slugs
    )
    page_groups = []
    for index, configured_page in enumerate(BUDGET_PAGE_GROUPS):
        page_groups.append({
            **configured_page,
            'count': (
                page_counts.get(configured_page['slug'], 0)
                + (unassigned_count if index == 0 else 0)
            ),
        })

    page_by_slug = {page['slug']: page for page in page_groups}
    active_page = page_by_slug.get(request.GET.get('group'), page_groups[0])
    if active_page['slug'] == BUDGET_PAGE_GROUPS[0]['slug']:
        items = BudgetItem.objects.filter(
            Q(page_group=active_page['slug']) |
            ~Q(page_group__in=valid_page_slugs)
        )
    else:
        items = BudgetItem.objects.filter(page_group=active_page['slug'])
    if query:
        items = items.filter(
            Q(budget_code__icontains=query) |
            Q(category__icontains=query) |
            Q(ministry__icontains=query) |
            Q(lead_pastor__icontains=query) |
            Q(accounting_subject__icontains=query)
        )
    items = items.order_by('id')
    logs = BudgetChangeLog.objects.select_related('changed_by', 'budget_item').order_by('-created_at', '-id')[:200]
    return render(request, 'budget/budget_list.html', {
        'fields': FIELDS,
        'form_fields': FORM_FIELDS,
        'items': items,
        'page_groups': page_groups,
        'active_page': active_page,
        'query': query,
        'total_count': items.count(),
        'all_count': BudgetItem.objects.count(),
        'logs': logs,
    })


@login_required
def budget_edit(request, pk):
    item = get_object_or_404(BudgetItem, pk=pk)
    if request.method == 'POST':
        before = _snapshot(item)
        for field, value in _payload_from_post(request.POST, item.page_group).items():
            setattr(item, field, value)
        item.save()
        _log_change(request, 'update', item, before, _snapshot(item))
    return redirect('budget-list')


@login_required
def budget_delete(request, pk):
    item = get_object_or_404(BudgetItem, pk=pk)
    if request.method == 'POST':
        before = _snapshot(item)
        _log_change(request, 'delete', item, before, None)
        item.delete()
    return redirect(request.POST.get('next') or 'budget-list')


@login_required
def budget_export(request):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = '預算表維護'
    sheet.append([label for _field, label in FORM_FIELDS] + ['所屬分頁', '使用比例', '餘額'])
    for item in BudgetItem.objects.order_by('id'):
        ratio = item.usage_ratio
        sheet.append([
            item.category,
            item.budget_code,
            item.ministry,
            item.annual_goal,
            item.strategy_plan,
            item.activity_budget,
            item.lead_pastor,
            float(item.budget_2026) if item.budget_2026 is not None else None,
            float(item.used_amount) if item.used_amount is not None else None,
            item.accounting_subject,
            item.get_page_group_display(),
            float(ratio) / 100 if ratio is not None else None,
            float(item.balance),
        ])
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="budget_items_2026.xlsx"'
    return response
