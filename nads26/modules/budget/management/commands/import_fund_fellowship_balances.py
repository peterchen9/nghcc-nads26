import os
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction
from modules.budget.models import Fund, FundBalance, Fellowship, FellowshipBalance
from modules.menu.models import MenuItem


class Command(BaseCommand):
    help = 'Import fund and fellowship balances from 基金與團契款餘額.xlsx'

    def handle(self, *args, **options):
        file_path = '基金與團契款餘額.xlsx'
        if not os.path.exists(file_path):
            self.stderr.write(f'File not found: {file_path}')
            return

        self.stdout.write(f'Loading workbook: {file_path} ...')
        wb = openpyxl.load_workbook(file_path, data_only=True)

        # Clear existing records to ensure idempotency
        self.stdout.write('Clearing existing funds, fellowships, and their balances...')
        with transaction.atomic():
            FundBalance.objects.all().delete()
            Fund.objects.all().delete()
            FellowshipBalance.objects.all().delete()
            Fellowship.objects.all().delete()

        # 1. Parse Sheet "基金餘額"
        if "基金餘額" in wb.sheetnames:
            self.stdout.write('Parsing sheet "基金餘額" ...')
            s_fund = wb['基金餘額']

            # Map column to Fund
            funds_map = {}
            for col in range(2, s_fund.max_column + 1):
                name = s_fund.cell(1, col).value
                if name and str(name).strip() not in ('合計', 'Total', '總計', '合計總額'):
                    name = str(name).strip()
                    note_val = s_fund.cell(2, col).value
                    note = str(note_val).strip() if note_val else ''

                    fund = Fund.objects.create(
                        name=name,
                        note=note,
                        order=col
                    )
                    funds_map[col] = fund
                    self.stdout.write(f'Created Fund: {fund.name} (col={col})')

            # Insert monthly balances
            fund_balances_to_create = []
            for r in range(3, s_fund.max_row + 1):
                date_val = s_fund.cell(r, 1).value
                if date_val is None:
                    continue

                # Parse date
                dt = self.parse_date(date_val)
                if not dt:
                    continue

                for col, fund in funds_map.items():
                    val = s_fund.cell(r, col).value
                    if val is not None and val != '':
                        try:
                            if isinstance(val, str):
                                val = val.replace(',', '').replace('$', '').strip()
                            balance_dec = Decimal(str(val))
                            fund_balances_to_create.append(FundBalance(
                                fund=fund,
                                month=dt,
                                balance=balance_dec
                            ))
                        except (InvalidOperation, ValueError):
                            pass

            if fund_balances_to_create:
                with transaction.atomic():
                    FundBalance.objects.bulk_create(fund_balances_to_create, batch_size=500)
                self.stdout.write(f'Successfully imported {len(fund_balances_to_create)} fund balance records.')

        # 2. Parse Sheet "團契餘額"
        if "團契餘額" in wb.sheetnames:
            self.stdout.write('Parsing sheet "團契餘額" ...')
            s_fellowship = wb['團契餘額']

            # Map column to Fellowship
            fellowships_map = {}
            for col in range(2, s_fellowship.max_column + 1):
                name = s_fellowship.cell(1, col).value
                if name and str(name).strip() not in ('合計', 'Total', '總計', '合計總額'):
                    name = str(name).strip()
                    note_val = s_fellowship.cell(2, col).value
                    note = str(note_val).strip() if note_val else ''

                    fellowship = Fellowship.objects.create(
                        name=name,
                        note=note,
                        order=col
                    )
                    fellowships_map[col] = fellowship
                    self.stdout.write(f'Created Fellowship: {fellowship.name} (col={col})')

            # Insert monthly balances
            fellowship_balances_to_create = []
            for r in range(3, s_fellowship.max_row + 1):
                date_val = s_fellowship.cell(r, 1).value
                if date_val is None:
                    continue

                dt = self.parse_date(date_val)
                if not dt:
                    continue

                for col, fellowship in fellowships_map.items():
                    val = s_fellowship.cell(r, col).value
                    if val is not None and val != '':
                        try:
                            if isinstance(val, str):
                                val = val.replace(',', '').replace('$', '').strip()
                            balance_dec = Decimal(str(val))
                            fellowship_balances_to_create.append(FellowshipBalance(
                                fellowship=fellowship,
                                month=dt,
                                balance=balance_dec
                            ))
                        except (InvalidOperation, ValueError):
                            pass

            if fellowship_balances_to_create:
                with transaction.atomic():
                    FellowshipBalance.objects.bulk_create(fellowship_balances_to_create, batch_size=500)
                self.stdout.write(f'Successfully imported {len(fellowship_balances_to_create)} fellowship balance records.')

        # 3. Create or Update MenuItem under "財會"
        try:
            parent_menu = MenuItem.objects.filter(title="財會", parent=None).first()
            if parent_menu:
                menu_item, created = MenuItem.objects.get_or_create(
                    title="基金團契餘額",
                    parent=parent_menu,
                    defaults={
                        'route': '/finance/fund-fellowship-balances/',
                        'order': 130,
                        'roles': '*',
                        'is_active': True,
                    }
                )
                if not created:
                    menu_item.route = '/finance/fund-fellowship-balances/'
                    menu_item.save()
                self.stdout.write(self.style.SUCCESS(f'MenuItem "基金團契餘額" seeded successfully.'))
            else:
                self.stdout.write(self.style.WARNING('Parent menu "財會" not found. Skip MenuItem seeding.'))
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Error seeding MenuItem: {str(e)}'))

        self.stdout.write(self.style.SUCCESS('Fund and Fellowship balances import finished!'))

    def parse_date(self, date_val):
        if isinstance(date_val, datetime):
            return date_val.date()
        elif isinstance(date_val, date):
            return date_val
        elif isinstance(date_val, str):
            for fmt in ('%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%Y/%m/%d', '%Y/%m/%d %H:%M:%S'):
                try:
                    return datetime.strptime(date_val.strip(), fmt).date()
                except ValueError:
                    pass
        return None
