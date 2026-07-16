import os
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction
from modules.budget.models import MonthlyOffering, AnnualOffering
from modules.menu.models import MenuItem


class Command(BaseCommand):
    help = 'Import offering statistics from 奉獻金額與人數統計(密).xlsx'

    def handle(self, *args, **options):
        file_path = '奉獻金額與人數統計(密).xlsx'
        if not os.path.exists(file_path):
            self.stderr.write(f'File not found: {file_path}')
            return

        self.stdout.write(f'Loading workbook: {file_path} ...')
        wb = openpyxl.load_workbook(file_path, data_only=True)

        self.stdout.write('Clearing existing monthly and annual offering records...')
        with transaction.atomic():
            MonthlyOffering.objects.all().delete()
            AnnualOffering.objects.all().delete()

        # 1. Parse Sheet "年統計"
        if "年統計" in wb.sheetnames:
            self.stdout.write('Parsing sheet "年統計" ...')
            ws_year = wb['年統計']
            annual_offerings_to_create = []

            for r in range(2, ws_year.max_row + 1):
                year_val = ws_year.cell(r, 1).value
                if year_val is None:
                    continue

                try:
                    year = int(float(str(year_val)))
                except (ValueError, TypeError):
                    continue

                # Core data: cols 2 to 18
                pledge_people = self.clean_int(ws_year.cell(r, 2).value)
                
                sunday = self.clean_decimal(ws_year.cell(r, 3).value)
                pledge = self.clean_decimal(ws_year.cell(r, 4).value)
                thanksgiving = self.clean_decimal(ws_year.cell(r, 5).value)
                building = self.clean_decimal(ws_year.cell(r, 6).value)
                planting = self.clean_decimal(ws_year.cell(r, 7).value)
                designated = self.clean_decimal(ws_year.cell(r, 8).value)
                mission = self.clean_decimal(ws_year.cell(r, 9).value)
                galilee = self.clean_decimal(ws_year.cell(r, 10).value)
                timothy = self.clean_decimal(ws_year.cell(r, 11).value)
                kingdom = self.clean_decimal(ws_year.cell(r, 12).value)
                venue = self.clean_decimal(ws_year.cell(r, 13).value)
                spring = self.clean_decimal(ws_year.cell(r, 14).value)
                other = self.clean_decimal(ws_year.cell(r, 15).value)
                project = self.clean_decimal(ws_year.cell(r, 16).value)
                total = self.clean_decimal(ws_year.cell(r, 17).value)

                # Skip blank/empty records (year placeholder rows at the end)
                if not any([pledge_people, sunday, pledge, thanksgiving, building, planting, designated, mission, galilee, timothy, kingdom, venue, spring, other, project]):
                    continue

                annual_offerings_to_create.append(AnnualOffering(
                    year=year,
                    pledge_people_count=pledge_people,
                    sunday=sunday,
                    pledge=pledge,
                    thanksgiving=thanksgiving,
                    building=building,
                    planting=planting,
                    designated=designated,
                    mission=mission,
                    galilee=galilee,
                    timothy=timothy,
                    kingdom=kingdom,
                    venue=venue,
                    spring=spring,
                    other=other,
                    project=project,
                    total=total
                ))
                self.stdout.write(f'Read Year: {year}')

            if annual_offerings_to_create:
                with transaction.atomic():
                    AnnualOffering.objects.bulk_create(annual_offerings_to_create)
                self.stdout.write(f'Successfully imported {len(annual_offerings_to_create)} annual offering records.')

        # 2. Parse Sheet "月統計"
        if "月統計" in wb.sheetnames:
            self.stdout.write('Parsing sheet "月統計" ...')
            ws_month = wb['月統計']
            monthly_offerings_to_create = []

            for r in range(2, ws_month.max_row + 1):
                date_cell = ws_month.cell(r, 1).value
                if date_cell is None:
                    continue

                month_date = self.parse_month(date_cell)
                if not month_date:
                    continue

                pledge_people = self.clean_int(ws_month.cell(r, 2).value)
                
                sunday = self.clean_decimal(ws_month.cell(r, 3).value)
                pledge = self.clean_decimal(ws_month.cell(r, 4).value)
                thanksgiving = self.clean_decimal(ws_month.cell(r, 5).value)
                building = self.clean_decimal(ws_month.cell(r, 6).value)
                planting = self.clean_decimal(ws_month.cell(r, 7).value)
                designated = self.clean_decimal(ws_month.cell(r, 8).value)
                mission = self.clean_decimal(ws_month.cell(r, 9).value)
                galilee = self.clean_decimal(ws_month.cell(r, 10).value)
                timothy = self.clean_decimal(ws_month.cell(r, 11).value)
                kingdom = self.clean_decimal(ws_month.cell(r, 12).value)
                venue = self.clean_decimal(ws_month.cell(r, 13).value)
                total = self.clean_decimal(ws_month.cell(r, 14).value)

                monthly_offerings_to_create.append(MonthlyOffering(
                    month=month_date,
                    pledge_people_count=pledge_people,
                    sunday=sunday,
                    pledge=pledge,
                    thanksgiving=thanksgiving,
                    building=building,
                    planting=planting,
                    designated=designated,
                    mission=mission,
                    galilee=galilee,
                    timothy=timothy,
                    kingdom=kingdom,
                    venue=venue,
                    total=total
                ))
                self.stdout.write(f'Read Month: {month_date.strftime("%Y-%m")}')

            if monthly_offerings_to_create:
                with transaction.atomic():
                    MonthlyOffering.objects.bulk_create(monthly_offerings_to_create)
                self.stdout.write(f'Successfully imported {len(monthly_offerings_to_create)} monthly offering records.')

        # 3. Create/Update MenuItem under "財會"
        try:
            parent_menu = MenuItem.objects.filter(title="財會", parent=None).first()
            if parent_menu:
                menu_item, created = MenuItem.objects.get_or_create(
                    title="奉獻統計",
                    parent=parent_menu,
                    defaults={
                        'route': '/finance/offering-statistics/',
                        'order': 140,
                        'roles': '*',
                        'is_active': True,
                    }
                )
                if not created:
                    menu_item.route = '/finance/offering-statistics/'
                    menu_item.save()
                self.stdout.write(self.style.SUCCESS('MenuItem "奉獻統計" seeded successfully.'))
            else:
                self.stdout.write(self.style.WARNING('Parent menu "財會" not found. Skip MenuItem seeding.'))
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Error seeding MenuItem: {str(e)}'))

        self.stdout.write(self.style.SUCCESS('Offering statistics import finished!'))

    def parse_month(self, val):
        if isinstance(val, (datetime, date)):
            return date(val.year, val.month, 1)
        elif isinstance(val, str):
            val = val.strip()
            for sep in ('-', '/'):
                if sep in val:
                    parts = val.split(sep)
                    if len(parts) >= 2:
                        try:
                            y = int(parts[0])
                            m = int(parts[1])
                            return date(y, m, 1)
                        except ValueError:
                            pass
        return None

    def clean_decimal(self, val):
        if val is None or val == '':
            return Decimal('0.0')
        if isinstance(val, str):
            val = val.replace(',', '').replace('$', '').strip()
        try:
            return Decimal(str(val))
        except (InvalidOperation, ValueError):
            return Decimal('0.0')

    def clean_int(self, val):
        if val is None or val == '':
            return None
        try:
            return int(float(str(val)))
        except (ValueError, TypeError):
            return None
