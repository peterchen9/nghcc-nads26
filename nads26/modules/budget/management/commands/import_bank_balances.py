import os
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction
from modules.budget.models import BankAccount, AccountBalance, TimeDeposit


class Command(BaseCommand):
    help = 'Import bank accounts, balances, and time deposits from 銀行帳戶結餘表.xlsx'

    def handle(self, *args, **options):
        file_path = '銀行帳戶結餘表.xlsx'
        if not os.path.exists(file_path):
            self.stderr.write(f'File not found: {file_path}')
            return

        self.stdout.write(f'Loading workbook: {file_path} ...')
        wb = openpyxl.load_workbook(file_path, data_only=True)

        # 1. Clear existing records to ensure idempotency
        self.stdout.write('Clearing existing bank accounts, balances, and time deposits...')
        with transaction.atomic():
            AccountBalance.objects.all().delete()
            BankAccount.objects.all().delete()
            TimeDeposit.objects.all().delete()

        # 2. Parse Sheet 1: 銀行帳戶結餘
        self.stdout.write('Parsing sheet "銀行帳戶結餘" ...')
        s1 = wb['銀行帳戶結餘']
        
        # Identify bank account columns
        accounts_map = {} # col_idx -> BankAccount object
        for col in range(2, s1.max_column + 1):
            category = s1.cell(2, col).value
            # Check if it is a valid bank account column
            if category and str(category).strip() not in ('合計', 'Total', '總計', '合計總額'):
                bank = s1.cell(3, col).value or ''
                acc_no = s1.cell(4, col).value or ''
                
                # Clean up values
                category = str(category).strip()
                bank = str(bank).strip() if bank else ''
                acc_no = str(acc_no).strip() if acc_no else ''
                
                account = BankAccount.objects.create(
                    category=category,
                    bank=bank,
                    account_no=acc_no,
                    order=col
                )
                accounts_map[col] = account
                self.stdout.write(f'Created BankAccount: {account} (col={col})')

        # Insert monthly balances
        balances_to_create = []
        for r in range(5, s1.max_row + 1):
            date_val = s1.cell(r, 1).value
            if date_val is None:
                continue
                
            # Parse date
            if isinstance(date_val, datetime):
                dt = date_val.date()
            elif isinstance(date_val, date):
                dt = date_val
            elif isinstance(date_val, str):
                try:
                    dt = datetime.strptime(date_val.strip(), '%Y-%m-%d').date()
                except ValueError:
                    try:
                        dt = datetime.strptime(date_val.strip(), '%Y-%m-%d %H:%M:%S').date()
                    except ValueError:
                        continue
            else:
                continue

            # Check if this row actually has any non-None balances
            row_has_balances = False
            for col, account in accounts_map.items():
                val = s1.cell(r, col).value
                if val is not None and val != '':
                    row_has_balances = True
                    break
            
            if not row_has_balances:
                continue

            for col, account in accounts_map.items():
                val = s1.cell(r, col).value
                if val is not None and val != '':
                    try:
                        # Strip formatting if string
                        if isinstance(val, str):
                            val = val.replace(',', '').replace('$', '').strip()
                        balance_dec = Decimal(str(val))
                        balances_to_create.append(AccountBalance(
                            account=account,
                            month=dt,
                            balance=balance_dec
                        ))
                    except (InvalidOperation, ValueError):
                        pass

        if balances_to_create:
            with transaction.atomic():
                AccountBalance.objects.bulk_create(balances_to_create, batch_size=500)
            self.stdout.write(f'Successfully imported {len(balances_to_create)} monthly balance records.')

        # 3. Parse Sheet 2: 定期存單
        self.stdout.write('Parsing sheet "定期存單" ...')
        s2 = wb['定期存單']
        
        current_cat = None
        deposits_to_create = []
        order_counter = 0
        
        for r in range(2, s2.max_row + 1):
            row_vals = [s2.cell(r, c).value for c in range(1, 9)]
            if not any(row_vals[1:8]): # If all columns except maybe category are empty, skip
                continue
                
            cat = row_vals[0]
            if cat is not None:
                current_cat = str(cat).strip()
            
            if not current_cat:
                continue
                
            duration = str(row_vals[1]).strip() if row_vals[1] is not None else ''
            period = str(row_vals[2]).strip() if row_vals[2] is not None else ''
            dep_type = str(row_vals[3]).strip() if row_vals[3] is not None else ''
            dep_no = str(row_vals[4]).strip() if row_vals[4] is not None else ''
            
            rate_val = row_vals[5]
            interest_rate = None
            if rate_val is not None:
                try:
                    interest_rate = Decimal(str(rate_val))
                except (InvalidOperation, ValueError):
                    pass
                    
            rate_type = str(row_vals[6]).strip() if row_vals[6] is not None else ''
            
            amount_val = row_vals[7]
            if amount_val is None:
                continue
            try:
                if isinstance(amount_val, str):
                    amount_val = amount_val.replace(',', '').replace('$', '').strip()
                amount_dec = Decimal(str(amount_val))
            except (InvalidOperation, ValueError):
                continue
                
            order_counter += 1
            deposits_to_create.append(TimeDeposit(
                category=current_cat,
                duration=duration,
                period=period,
                deposit_type=dep_type,
                deposit_no=dep_no,
                interest_rate=interest_rate,
                rate_type=rate_type,
                amount=amount_dec,
                order=order_counter
            ))

        if deposits_to_create:
            with transaction.atomic():
                TimeDeposit.objects.bulk_create(deposits_to_create)
            self.stdout.write(f'Successfully imported {len(deposits_to_create)} time deposit records.')
            
        self.stdout.write(self.style.SUCCESS('Bank accounts, balances, and time deposits import finished!'))
