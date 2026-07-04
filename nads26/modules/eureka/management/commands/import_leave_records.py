import os
import re
import datetime
import requests
import openpyxl
from django.core.management.base import BaseCommand
from django.conf import settings
from django.db import transaction
from modules.eureka.models import StaffLeave, DailyDutyNote, Member, StaffInfo
from modules.accounts.models import SystemSetting

STAFF_NAME_MAP = {
    '明月': '林明月',
    '德官': '董德官',
    '明珠': '羅明珠',
    '玉筍': '周玉筍',
    '宜庭': '何宜庭',
    '仲甫': '鄭仲甫',
    '慕聖': '張慕聖',
    '沐恩': '趙沐恩',
    '囿余': '陳囿余',
    '小慧': '謝淑慧',
    '欣娣': '吳欣娣',
    '文正': '林文正',
    '美美': '黃美美',
    '惠萍': '張惠萍',
    '文秀': '蔡文秀',
    '依蓮': '陳依蓮',
    '宗英': '楊宗英',
    '玉書': '林玉書',
    '方正': '施方正',
    '開門': '王文村', # Handled mapping if needed
    '慧芝': '郭慧芝',
    '美芳': '林美芳',
    '潘傳': '陳潘傳',
}

def get_member_for_name(full_name):
    if full_name == '林美芳':
        return Member.objects.filter(church_id=7524).first()
    return Member.objects.filter(name=full_name).first()

class Command(BaseCommand):
    help = 'Import staff leave records and staff info from Google Sheets'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            dest='file_path',
            help='Import leave records from a local .xlsx file instead of downloading configured sheets.',
        )

    def handle(self, *args, **options):
        self.stdout.write("Starting leave records and staff info import...")

        file_path = options.get('file_path')
        if file_path:
            if not os.path.exists(file_path):
                self.stderr.write(self.style.ERROR(f"File not found: {file_path}"))
                return

            try:
                self.parse_xlsx(file_path)
                self.parse_personnel(file_path)
            except Exception as e:
                self.stderr.write(self.style.ERROR(f"Error parsing {file_path}: {e}"))
                raise

            self.stdout.write(self.style.SUCCESS("Leave records and staff info import complete!"))
            return

        urls = {}
        settings_qs = SystemSetting.objects.filter(key__startswith='VACATION_SHEET_')
        for s in settings_qs:
            year_match = re.search(r'\d{4}', s.key)
            if year_match:
                urls[year_match.group()] = s.value.strip()

        if not urls:
            self.stdout.write("No SystemSetting starting with 'VACATION_SHEET_' found. Using default URLs for 2024 and 2026.")
            urls = {
                "2026": "https://docs.google.com/spreadsheets/d/1w5qG4sanVTaQb405a5Gyq2g1_6CKViRp/export?format=xlsx",
                "2024": "https://docs.google.com/spreadsheets/d/1IljkEh2k4aOeFsTU-UpUKSvn20TR4kwp/export?format=xlsx"
            }

        tmp_dir = os.path.join(settings.MEDIA_ROOT, 'tmp')
        os.makedirs(tmp_dir, exist_ok=True)

        for year, url in urls.items():
            if "/edit" in url:
                url = url.split("/edit")[0] + "/export?format=xlsx"

            dest_path = os.path.join(tmp_dir, f"vacation_{year}.xlsx")
            self.stdout.write(f"Downloading sheet for year {year} from: {url} ...")
            
            try:
                response = requests.get(url, timeout=30)
                response.raise_for_status()
                with open(dest_path, "wb") as f:
                    f.write(response.content)
                self.stdout.write(self.style.SUCCESS(f"Successfully downloaded to {dest_path}"))
            except Exception as e:
                self.stderr.write(self.style.ERROR(f"Failed to download sheet for year {year}: {e}"))
                continue

            try:
                self.parse_xlsx(dest_path)
                self.parse_personnel(dest_path)
            except Exception as e:
                self.stderr.write(self.style.ERROR(f"Error parsing {dest_path}: {e}"))

        self.stdout.write(self.style.SUCCESS("Leave records and staff info import complete!"))

    def parse_xlsx(self, filepath):
        wb = openpyxl.load_workbook(filepath, data_only=True)
        pattern = re.compile(r'^\d{4}-\d{2}$')
        
        for sheetname in wb.sheetnames:
            if not pattern.match(sheetname):
                continue
                
            self.stdout.write(f"Parsing worksheet: {sheetname}")
            ws = wb[sheetname]
            
            header_row_idx = None
            rows = list(ws.iter_rows(values_only=True))
            for r_idx, row in enumerate(rows[:-1]):
                row_values = [str(v).strip() if v is not None else "" for v in row]
                next_values = [str(v).strip() if v is not None else "" for v in rows[r_idx + 1]]

                has_calendar_labels = (
                    any(v in ('日期', '日', '星期') for v in row_values[:3])
                    or ('星期' in row_values[:3] and '教會行政' in row_values[:3])
                )
                has_slot_labels = any(v in ('上', '下', 'AM', 'PM', '上午', '下午') for v in next_values[3:])

                if has_calendar_labels and has_slot_labels:
                    header_row_idx = r_idx
                    break
                    
            if header_row_idx is None:
                self.stdout.write(f"  Header row not found in worksheet {sheetname}, skipping.")
                continue

            name_row = rows[header_row_idx]
            sub_row = rows[header_row_idx + 1] if header_row_idx + 1 < len(rows) else []
            
            year_val, month_val = map(int, sheetname.split('-'))
            
            staff_columns = []
            for c_idx in range(3, len(name_row)):
                name = name_row[c_idx]
                if name and str(name).strip() not in ('星期', '教會行政', '星期 ', ''):
                    clean_name = str(name).strip()
                    am_col = c_idx
                    pm_col = None
                    next_slot = str(sub_row[c_idx + 1]).strip() if c_idx + 1 < len(sub_row) and sub_row[c_idx + 1] is not None else ""
                    if next_slot in ('下', 'PM', '下午'):
                        pm_col = c_idx + 1
                    
                    staff_columns.append({
                        'short_name': clean_name,
                        'am_col': am_col,
                        'pm_col': pm_col
                    })

            self.stdout.write(f"  Found {len(staff_columns)} staff columns. Writing to database in a single transaction...")
            
            with transaction.atomic():
                for r_idx in range(header_row_idx + 2, len(rows)):
                    row = rows[r_idx]
                    if not row or not any(row):
                        continue
                        
                    day_val = row[0]
                    if day_val is None:
                        continue
                        
                    day = None
                    if isinstance(day_val, datetime.datetime):
                        day = day_val.day
                    elif isinstance(day_val, int):
                        day = day_val
                    elif isinstance(day_val, str) and day_val.strip().isdigit():
                        day = int(day_val.strip())
                    else:
                        try:
                            day = int(float(str(day_val).strip()))
                        except ValueError:
                            pass
                            
                    if day is None or not (1 <= day <= 31):
                        continue

                    try:
                        date = datetime.date(year_val, month_val, day)
                    except ValueError:
                        continue

                    # 1. Save Admin/Event Note
                    admin_note = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
                    if admin_note and admin_note != "None":
                        DailyDutyNote.objects.update_or_create(
                            date=date,
                            defaults={'note': admin_note}
                        )

                    # 2. Save Staff Leaves
                    for sc in staff_columns:
                        short_name = sc['short_name']
                        full_name = STAFF_NAME_MAP.get(short_name, short_name)
                        member = get_member_for_name(full_name)
                        
                        am_val = str(row[sc['am_col']]).strip() if sc['am_col'] < len(row) and row[sc['am_col']] is not None else ""
                        pm_val = str(row[sc['pm_col']]).strip() if sc['pm_col'] is not None and sc['pm_col'] < len(row) and row[sc['pm_col']] is not None else ""

                        if am_val == "None": am_val = ""
                        if pm_val == "None": pm_val = ""

                        # AM
                        if am_val:
                            StaffLeave.objects.update_or_create(
                                staff_name=short_name,
                                date=date,
                                time_slot='AM',
                                defaults={
                                    'member': member,
                                    'leave_type': am_val
                                }
                            )
                        else:
                            StaffLeave.objects.filter(staff_name=short_name, date=date, time_slot='AM').delete()

                        # PM
                        if pm_val:
                            StaffLeave.objects.update_or_create(
                                staff_name=short_name,
                                date=date,
                                time_slot='PM',
                                defaults={
                                    'member': member,
                                    'leave_type': pm_val
                                }
                            )
                        else:
                            StaffLeave.objects.filter(staff_name=short_name, date=date, time_slot='PM').delete()

    def parse_personnel(self, filepath):
        self.stdout.write("Parsing personnel sheet (人員)...")
        wb = openpyxl.load_workbook(filepath, data_only=True)
        if "人員" not in wb.sheetnames:
            self.stdout.write("  Worksheet '人員' not found.")
            return
            
        ws = wb["人員"]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return
            
        # Headers are in row 0
        headers = rows[0]
        year_cols = {}
        for c_idx, h in enumerate(headers):
            if h is not None:
                try:
                    val = float(h)
                    if 2000 <= val <= 2100:
                        year_cols[str(int(val))] = c_idx
                except ValueError:
                    pass

        self.stdout.write(f"  Found leave quota year columns: {', '.join(sorted(year_cols.keys()))}")

        with transaction.atomic():
            for r_idx in range(1, len(rows)):
                row = rows[r_idx]
                if not row or row[0] is None:
                    continue
                    
                try:
                    staff_id = int(float(row[0]))
                except ValueError:
                    continue
                    
                name = str(row[1]).strip() if row[1] is not None else ""
                if not name:
                    continue
                    
                identity_code = str(row[2]).strip() if row[2] is not None else ""
                
                is_active = True
                if row[3] is not None:
                    try:
                        active_val = float(row[3])
                        is_active = (active_val == 1.0)
                    except ValueError:
                        pass
                        
                email = str(row[4]).strip() if row[4] is not None else ""
                if email == "None" or email.lower() == "none": email = ""
                cc_email = str(row[5]).strip() if row[5] is not None else ""
                if cc_email == "None" or cc_email.lower() == "none": cc_email = ""
                
                onboard_date = None
                onboard_val = row[6]
                if isinstance(onboard_val, datetime.datetime):
                    onboard_date = onboard_val.date()
                elif isinstance(onboard_val, datetime.date):
                    onboard_date = onboard_val
                elif isinstance(onboard_val, str) and onboard_val.strip():
                    try:
                        onboard_date = datetime.datetime.strptime(onboard_val.strip(), "%Y-%m-%d").date()
                    except ValueError:
                        pass
                        
                quotas = {}
                for yr, c_idx in year_cols.items():
                    val = row[c_idx]
                    if val is not None:
                        try:
                            quotas[yr] = float(val)
                        except ValueError:
                            pass
                            
                StaffInfo.objects.update_or_create(
                    staff_id=staff_id,
                    defaults={
                        'name': name,
                        'identity_code': identity_code,
                        'is_active': is_active,
                        'email': email,
                        'cc_email': cc_email,
                        'onboard_date': onboard_date,
                        'leave_quotas': quotas
                    }
                )
        self.stdout.write(self.style.SUCCESS("  Personnel sheet parsing completed successfully!"))
