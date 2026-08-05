import os
import datetime
import openpyxl
from django.core.management.base import BaseCommand
from modules.eureka.models import YearlyAttendance, WeeklyAttendance
from modules.menu.models import MenuItem

class Command(BaseCommand):
    help = 'Import worship attendance data from Excel file and setup menu item'

    def handle(self, *args, **options):
        self.stdout.write("Starting worship attendance import...")
        
        file_path = "WS-主日聚會 (1).xlsx"
        if not os.path.exists(file_path):
            self.stderr.write(self.style.ERROR(f"Excel file not found at: {file_path}"))
            return

        # Load Excel with data_only=True to evaluate formulas like =A2+7
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            self.stderr.write(self.style.ERROR(f"Failed to load Excel file: {e}"))
            return

        # 1. Import YearlyAttendance (年統計)
        if "年統計" in wb.sheetnames:
            self.stdout.write("Importing yearly statistics...")
            sheet_years = wb["年統計"]
            rows = list(sheet_years.iter_rows(values_only=True))
            
            # First row is header: ('年', '聚會人數', '受洗', '備註')
            imported_count = 0
            for row in rows[1:]:
                year_val = row[0]
                if year_val is None:
                    continue
                try:
                    year = int(year_val)
                except (TypeError, ValueError):
                    self.stdout.write(f"Skipping invalid year row: {row}")
                    continue

                attendance = row[1] if row[1] is not None else 0
                baptized = row[2] if row[2] is not None else 0
                remark = str(row[3]).strip() if row[3] is not None else ""

                YearlyAttendance.objects.update_or_create(
                    year=year,
                    defaults={
                        'attendance': attendance,
                        'baptized': baptized,
                        'remark': remark
                    }
                )
                imported_count += 1
            self.stdout.write(self.style.SUCCESS(f"Imported {imported_count} yearly statistics rows."))
        else:
            self.stderr.write(self.style.WARNING("Sheet '年統計' not found in Excel."))

        # 2. Import WeeklyAttendance (每週)
        if "每週" in wb.sheetnames:
            self.stdout.write("Importing weekly statistics...")
            sheet_weeks = wb["每週"]
            rows = list(sheet_weeks.iter_rows(values_only=True))
            
            # First row is header: ('週(主日日期)', '第一堂', '第二堂', '晚堂', '線上', '兒主', '青少', '總計', '主日新朋友', '註記(不含新朋友)')
            imported_count = 0
            for row in rows[1:]:
                date_val = row[0]
                if date_val is None:
                    continue
                
                # Parse date
                actual_date = None
                if isinstance(date_val, (datetime.datetime, datetime.date)):
                    if isinstance(date_val, datetime.datetime):
                        actual_date = date_val.date()
                    else:
                        actual_date = date_val
                elif isinstance(date_val, str):
                    date_str = date_val.strip()
                    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
                        try:
                            actual_date = datetime.datetime.strptime(date_str, fmt).date()
                            break
                        except ValueError:
                            continue
                
                if not actual_date:
                    self.stdout.write(f"Skipping row with invalid date: {row}")
                    continue

                first_service = row[1] if row[1] is not None else None
                second_service = row[2] if row[2] is not None else None
                evening_service = row[3] if row[3] is not None else None
                online = row[4] if row[4] is not None else None
                children = row[5] if row[5] is not None else None
                youth = row[6] if row[6] is not None else None
                total = row[7] if row[7] is not None else None
                new_friends = row[8] if row[8] is not None else None
                remark = str(row[9]).strip() if row[9] is not None else ""

                WeeklyAttendance.objects.update_or_create(
                    date=actual_date,
                    defaults={
                        'first_service': first_service,
                        'second_service': second_service,
                        'evening_service': evening_service,
                        'online': online,
                        'children': children,
                        'youth': youth,
                        'total': total,
                        'new_friends': new_friends,
                        'remark': remark
                    }
                )
                imported_count += 1
            self.stdout.write(self.style.SUCCESS(f"Imported {imported_count} weekly statistics rows."))
        else:
            self.stderr.write(self.style.WARNING("Sheet '每週' not found in Excel."))

        # 3. Create MenuItem for "聚會人數"
        parent_menu = MenuItem.objects.filter(title="關懷", parent=None).first()
        if parent_menu:
            menu_item, created = MenuItem.objects.get_or_create(
                title="聚會人數",
                route="/eureka/meeting-attendance/",
                parent=parent_menu,
                defaults={
                    'order': 6,
                    'is_active': True,
                    'roles': '*'
                }
            )
            if created:
                self.stdout.write(self.style.SUCCESS(f"Created menu item '聚會人數' (ID: {menu_item.id}) under '關懷'"))
            else:
                # Update it to make sure route and active status are correct
                menu_item.route = "/eureka/meeting-attendance/"
                menu_item.is_active = True
                menu_item.save()
                self.stdout.write(self.style.SUCCESS("Updated existing '聚會人數' menu item."))
        else:
            self.stderr.write(self.style.WARNING("Parent menu item '關懷' not found in database. Cannot create menu item."))

        self.stdout.write(self.style.SUCCESS("Worship attendance import and menu setup complete!"))
