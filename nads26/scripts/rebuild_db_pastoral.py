import os
import sys
import openpyxl
import django

# Setup django environment
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(BASE_DIR)
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'nads26.settings')
django.setup()

from modules.eureka.models import Member, PastoralOverseer, PastoralSection, PastoralGroup

def main():
    excel_path = os.path.join(BASE_DIR, '牧區小組.xlsx')
    print(f"Loading Excel file {excel_path}...")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    print("Clearing existing pastoral database entries...")
    PastoralGroup.objects.all().delete()
    PastoralSection.objects.all().delete()
    PastoralOverseer.objects.all().delete()
    
    current_overseer = None
    current_section = None
    
    overseers_count = 0
    sections_count = 0
    groups_count = 0
    
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 10)]
        if not any(vals):
            continue
            
        overseer_name = (vals[0] or '').strip()
        section_name = (vals[1] or '').strip()
        leader1 = (vals[2] or '').strip()
        leader2 = (vals[3] or '').strip()
        group_name = (vals[4] or '').strip()
        target = (vals[6] or '').strip()
        location = (vals[7] or '').strip()
        meeting_time = (vals[8] or '').strip()
        
        # Forward fill
        if overseer_name:
            current_overseer = overseer_name
        if section_name:
            current_section = section_name
            
        if not current_overseer:
            continue
            
        # Get or create Overseer
        overseer_obj, created = PastoralOverseer.objects.get_or_create(name=current_overseer)
        if created:
            overseers_count += 1
            
        # Get or create Section
        sec_name = current_section or '無牧區'
        section_obj, created = PastoralSection.objects.get_or_create(
            name=sec_name,
            defaults={
                'overseer': overseer_obj,
                'leader': leader1,
                'counselor': leader2
            }
        )
        if created:
            sections_count += 1
        else:
            # If existed, update fields if defined in this row
            if leader1 and not section_obj.leader:
                section_obj.leader = leader1
            if leader2 and not section_obj.counselor:
                section_obj.counselor = leader2
            section_obj.save()
            
        # Get or create Group
        if group_name:
            group_obj, created = PastoralGroup.objects.get_or_create(
                name=group_name,
                defaults={
                    'section': section_obj,
                    'meeting_time': meeting_time,
                    'location': location,
                    'target': target
                }
            )
            if created:
                groups_count += 1
            else:
                # Update attributes if they were empty
                if meeting_time and not group_obj.meeting_time:
                    group_obj.meeting_time = meeting_time
                if location and not group_obj.location:
                    group_obj.location = location
                if target and not group_obj.target:
                    group_obj.target = target
                group_obj.save()
                
    print(f"Pastoral structure populated successfully:")
    print(f"  Overseers: {overseers_count}")
    print(f"  Sections: {sections_count}")
    print(f"  Groups: {groups_count}")
    
    print("Clearing section and family1 fields in Member database table...")
    member_count = Member.objects.all().update(section='', family1='')
    print(f"Cleared organization assignments for {member_count} members.")

if __name__ == '__main__':
    main()
