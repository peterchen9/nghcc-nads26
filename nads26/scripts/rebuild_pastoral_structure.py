import os
import sys
import openpyxl
import re
import django

# Setup django environment
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(BASE_DIR)
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'nads26.settings')
django.setup()

from modules.eureka.models import Member

def parse_excel():
    excel_path = os.path.join(BASE_DIR, '牧區小組.xlsx')
    print(f"Loading Excel file {excel_path}...")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    structure = {}
    current_overseer = None
    current_section = None
    
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 10)]
        if not any(vals):
            continue
            
        overseer = (vals[0] or '').strip()
        section = (vals[1] or '').strip()
        leader1 = (vals[2] or '').strip()
        leader2 = (vals[3] or '').strip()
        group_name = (vals[4] or '').strip()
        
        # Forward fill
        if overseer:
            current_overseer = overseer
        if section:
            current_section = section
            
        if not current_overseer:
            continue
            
        if current_overseer not in structure:
            structure[current_overseer] = {}
            
        sec_key = current_section or '無牧區'
        if sec_key not in structure[current_overseer]:
            structure[current_overseer][sec_key] = {
                'leader1': leader1,
                'leader2': leader2,
                'groups': set()
            }
            
        if leader1 and not structure[current_overseer][sec_key]['leader1']:
            structure[current_overseer][sec_key]['leader1'] = leader1
        if leader2 and not structure[current_overseer][sec_key]['leader2']:
            structure[current_overseer][sec_key]['leader2'] = leader2
            
        if group_name:
            structure[current_overseer][sec_key]['groups'].add(group_name)
            
    pastoral_structure = []
    for overseer, sections_dict in structure.items():
        sections_list = []
        for sec_name, sec_info in sections_dict.items():
            sections_list.append({
                'name': sec_name,
                'leader1': sec_info['leader1'],
                'leader2': sec_info['leader2'],
                'date': '',
                'status': '',
                'groups': sorted(list(sec_info['groups']))
            })
        pastoral_structure.append({
            'overseer': overseer,
            'sections': sections_list
        })
        
    return pastoral_structure

def rebuild_views():
    pastoral_structure = parse_excel()
    
    # Format pastoral_structure as a nice python list string
    import pprint
    structure_str = pprint.pformat(pastoral_structure, indent=4, width=120)
    
    views_path = os.path.join(BASE_DIR, 'modules', 'eureka', 'views.py')
    print(f"Reading views.py from {views_path}...")
    with open(views_path, 'r', encoding='utf-8') as f:
        content = f.read()
        
    # Replace pastoral_structure = [ ... ]
    # We find the start of pastoral_structure = [
    start_pattern = r'pastoral_structure\s*=\s*\['
    start_match = re.search(start_pattern, content)
    if not start_match:
        print("Error: Could not find pastoral_structure definition in views.py")
        return
        
    start_idx = start_match.start()
    
    # Find matching closing bracket for pastoral_structure = [
    bracket_count = 0
    end_idx = -1
    for i in range(start_idx, len(content)):
        if content[i] == '[':
            bracket_count += 1
        elif content[i] == ']':
            bracket_count -= 1
            if bracket_count == 0:
                end_idx = i + 1
                break
                
    if end_idx == -1:
        print("Error: Could not find closing bracket for pastoral_structure")
        return
        
    new_pastoral_structure_code = f"pastoral_structure = {structure_str}"
    content = content[:start_idx] + new_pastoral_structure_code + content[end_idx:]
    
    # Now, update the grouping logic around lines 368-376 (which is: groups_dict = {})
    old_group_logic = """            # 分組 (小組)
            groups_dict = {}
            for m in sec_members:
                g_name = m.family1
                if not g_name:
                    continue
                if g_name not in groups_dict:
                    groups_dict[g_name] = []
                groups_dict[g_name].append(m)"""
                
    new_group_logic = """            # 分組 (小組)
            groups_dict = {g: [] for g in sec.get('groups', [])}
            for m in sec_members:
                g_name = m.family1
                if not g_name:
                    continue
                if g_name not in groups_dict:
                    groups_dict[g_name] = []
                groups_dict[g_name].append(m)"""
                
    if old_group_logic in content:
        content = content.replace(old_group_logic, new_group_logic)
        print("Successfully updated grouping logic in views.py")
    else:
        # Try finding a slightly different spacing version
        normalized_content = content.replace('\r\n', '\n')
        normalized_old = old_group_logic.replace('\r\n', '\n')
        normalized_new = new_group_logic.replace('\r\n', '\n')
        if normalized_old in normalized_content:
            normalized_content = normalized_content.replace(normalized_old, normalized_new)
            content = normalized_content
            print("Successfully updated grouping logic in views.py (normalized)")
        else:
            print("Warning: Could not find grouping logic block to replace. Please check views.py manually.")
            
    print(f"Writing updated views.py...")
    with open(views_path, 'w', encoding='utf-8') as f:
        f.write(content)
    print("views.py updated successfully.")

def clear_db():
    print("Clearing section and family1 fields in Member database table...")
    # Clear section and family1 fields
    count = Member.objects.all().update(section='', family1='')
    print(f"Successfully cleared organization data for {count} members.")

if __name__ == '__main__':
    rebuild_views()
    clear_db()
